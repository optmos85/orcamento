"""
╔══════════════════════════════════════════════════════════════╗
║         ORÇAMENTO MENSAL — Backend FastAPI v4.0              ║
║  Instalar:                                                   ║
║    pip install fastapi uvicorn python-multipart              ║
║               chardet fpdf2 openpyxl                         ║
║  Rodar:   python app.py                                      ║
║  Acesse:  http://localhost:8000                              ║
╚══════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import json, os, io, re, hashlib, secrets, shutil, calendar as cal_mod
from pathlib import Path
from typing import Any, Optional
from datetime import datetime, date

import chardet, uvicorn
try:
    import requests as req_lib
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False
from fastapi import FastAPI, HTTPException, UploadFile, File, Query, Cookie, Response, Depends, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# ═══════════════════════ CONFIG ═══════════════════════════════

USERS_FILE  = Path("users.json")
BACKUP_DIR  = Path("backups")
STATIC_FILE = Path(__file__).parent / "index.html"
BACKUP_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Orçamento Mensal API", version="4.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"],
                   allow_headers=["*"], allow_credentials=True)

# Sessions em memória: token → user_id
_sessions: dict[str, str] = {}

# ═══════════════════════ DADOS PADRÃO ════════════════════════

DESPESAS_PADRAO = [
    ("Cartão de Crédito Nubank",   "Essencial"), ("Alimentação",         "Essencial"),
    ("Seguro Carro",               "Essencial"), ("Gasolina",            "Essencial"),
    ("Provisão IPVA",              "Essencial"), ("Conta celular",       "Essencial"),
    ("Curso",                      "Essencial"), ("Cartão de Crédito Inter",   "Essencial"),
    ("Cartão de Crédito Bradesco", "Essencial"), ("IPVA Parcelado",      "Essencial"),
    ("Notebook",                   "Essencial"), ("Consórcio",           "Poupança"),
    ("Poupança",                   "Poupança"),  ("Assinatura Globo Play","Desejo"),
    ("Netflix",                    "Desejo"),    ("Assinatura Programa X","Desejo"),
    ("Lazer / Passeio",            "Desejo"),    ("Aquisição Pessoal",   "Desejo"),
]

INTER_CATS = {
    "Alimentação":     ["supermercado","mercado","ifood","rappi","restaurante","lanchonete","padaria","hortifruti","carrefour","atacadão","assai"],
    "Gasolina":        ["posto","shell","ipiranga","petrobras","raizen","gasolina","etanol","combustivel"],
    "Conta celular":   ["tim","claro","vivo","oi ","nextel","celular"],
    "Netflix":         ["netflix"], "Assinatura Globo Play": ["globoplay","globo play"],
    "Lazer / Passeio": ["cinema","teatro","show","ingresso","spotify","deezer","disney+","hbo","apple tv"],
    "Poupança":        ["poupança","poupanca","investimento","cdb","tesouro","renda fixa"],
    "Cartão de Crédito Nubank":  ["nubank","nu pagamentos"],
    "Cartão de Crédito Inter":   ["inter pagamento","banco inter"],
    "Cartão de Crédito Bradesco":["bradesco"],
    "Curso":           ["udemy","coursera","alura","escola","faculdade","mensalidade"],
    "Seguro Carro":    ["seguro auto","seguro carro","porto seguro","azul seguro"],
    "Aquisição Pessoal":["amazon","shopee","mercado livre","americanas","submarino","magazine"],
}

PROFILE_COLORS = ["#2ecc85","#4f8fff","#ff9640","#a78bfa","#ff4f6e","#fbbf24","#22d3ee","#f472b6"]

# ═══════════════════════ AUTH ═════════════════════════════════

def hash_pw(password: str, salt: str) -> str:
    return hashlib.sha256(f"{salt}{password}".encode()).hexdigest()

def create_session(user_id: str) -> str:
    token = secrets.token_urlsafe(32)
    _sessions[token] = user_id
    return token

def get_session_user(session_token: Optional[str] = Cookie(default=None)) -> Optional[str]:
    if session_token and session_token in _sessions:
        return _sessions[session_token]
    return None

def require_auth(session_token: Optional[str] = Cookie(default=None)) -> str:
    user_id = get_session_user(session_token)
    if not user_id:
        raise HTTPException(401, "Não autenticado")
    return user_id

# ═══════════════════════ USERS ════════════════════════════════

def load_users() -> dict:
    if USERS_FILE.exists():
        return json.loads(USERS_FILE.read_text(encoding="utf-8"))
    return {"users": []}

def save_users(data: dict):
    USERS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def get_user(user_id: str) -> Optional[dict]:
    return next((u for u in load_users()["users"] if u["id"] == user_id), None)

def get_data_file(user_id: str) -> Path:
    return Path(f"data_{user_id}.json")

# ═══════════════════════ PERSISTÊNCIA ════════════════════════

def load_db(user_id: str) -> dict:
    f = get_data_file(user_id)
    # migração do arquivo legado
    if not f.exists() and Path("orcamento_data.json").exists():
        shutil.copy("orcamento_data.json", f)
    if f.exists():
        return json.loads(f.read_text(encoding="utf-8"))
    return {}

def save_db(user_id: str, data: dict):
    f = get_data_file(user_id)
    f.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    _auto_backup(user_id, data)

def _auto_backup(user_id: str, data: dict):
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = BACKUP_DIR / f"backup_{user_id}_{ts}.json"
    dst.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    # manter apenas últimos 30 backups por usuário
    baks = sorted(BACKUP_DIR.glob(f"backup_{user_id}_*.json"))
    for old in baks[:-30]: old.unlink()

def default_month() -> dict:
    return {
        "receitas": [
            {"nome": "Salário", "valor": 0.0, "data_credito": ""},
            {"nome": "Extra",   "valor": 0.0, "data_credito": ""},
        ],
        "despesas": [
            {"nome": n, "categoria": c, "prestacao": "", "valor": 0.0,
             "vencimento": "", "situacao": "", "data_pgto": "", "obs": "",
             "ordem": i, "recorrente": False}
            for i, (n, c) in enumerate(DESPESAS_PADRAO)
        ],
        "poupanca": {"saldo_anterior": 0.0, "aporte": 0.0, "juros": 0.0},
        "compras": [], "metas": [],
    }

def migrate(m: dict) -> dict:
    if isinstance(m.get("receitas"), dict):
        old = m["receitas"]
        m["receitas"] = [
            {"nome": "Salário", "valor": float(old.get("salario", 0)), "data_credito": ""},
            {"nome": "Extra",   "valor": float(old.get("extra",   0)), "data_credito": ""},
        ]
    for i, d in enumerate(m.get("despesas", [])):
        if "categoria"   not in d: d["categoria"]   = "Essencial"
        if "ordem"       not in d: d["ordem"]        = i
        if "recorrente"  not in d: d["recorrente"]   = False
    if "metas"   not in m: m["metas"]   = []
    if "compras" not in m: m["compras"] = []
    return m

def ensure_month(db: dict, key: str) -> dict:
    if key not in db:
        # busca recorrentes do mês mais recente
        prev_keys = sorted([k for k in db if not k.startswith("_") and k < key])
        if prev_keys:
            prev   = migrate(db[prev_keys[-1]])
            new_m  = default_month()
            new_m["despesas"] = [
                {**d, "valor": d["valor"] if d.get("recorrente") else 0.0,
                 "situacao": "", "data_pgto": "", "obs": ""}
                for d in prev["despesas"]
            ]
            db[key] = new_m
        else:
            db[key] = default_month()
    db[key] = migrate(db[key])
    return db

# ═══════════════════════ HELPERS ═════════════════════════════

def brl(v): return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")

def guess_cat(desc: str) -> str:
    low = desc.lower()
    for cat, kws in INTER_CATS.items():
        if any(k in low for k in kws): return cat
    return "Outros"

def clean_val(x: Any) -> float:
    s = str(x or "0").strip().replace("R$","").replace(" ","")
    if re.search(r"\d+\.\d{3},\d{2}$", s): s = s.replace(".","").replace(",",".")
    else:                                    s = s.replace(".","").replace(",",".")
    try: return float(s)
    except: return 0.0

def month_totals(m: dict):
    tr  = sum(r["valor"] for r in m["receitas"])
    td  = sum(d["valor"] for d in m["despesas"])
    ess = sum(d["valor"] for d in m["despesas"] if d.get("categoria") == "Essencial")
    pou = sum(d["valor"] for d in m["despesas"] if d.get("categoria") == "Poupança")
    des = sum(d["valor"] for d in m["despesas"] if d.get("categoria") == "Desejo")
    return tr, td, tr - td, ess, pou, des

def calc_score(m: dict, tr: float, td: float, ess: float, pou: float, des: float) -> dict:
    saldo = tr - td
    score = 0; details = []

    if saldo >= 0:
        score += 20; details.append({"item": "Saldo positivo", "pts": 20, "ok": True})
    else:
        details.append({"item": "Saldo negativo", "pts": 0, "ok": False})

    if tr > 0:
        pct_pou = pou / tr * 100
        if pct_pou >= 20: score += 20; details.append({"item": "Poupança ≥ 20%", "pts": 20, "ok": True})
        elif pct_pou >= 10: score += 10; details.append({"item": "Poupança ≥ 10%", "pts": 10, "ok": True})
        else: details.append({"item": f"Poupança {pct_pou:.0f}% (ideal ≥ 20%)", "pts": 0, "ok": False})

        pct_ess = ess / tr * 100
        if pct_ess <= 50: score += 20; details.append({"item": "Essenciais ≤ 50%", "pts": 20, "ok": True})
        elif pct_ess <= 60: score += 10; details.append({"item": f"Essenciais {pct_ess:.0f}%", "pts": 10, "ok": True})
        else: details.append({"item": f"Essenciais {pct_ess:.0f}% (ideal ≤ 50%)", "pts": 0, "ok": False})

        pct_des = des / tr * 100
        if pct_des <= 30: score += 15; details.append({"item": "Desejos ≤ 30%", "pts": 15, "ok": True})
        else: details.append({"item": f"Desejos {pct_des:.0f}% (ideal ≤ 30%)", "pts": 0, "ok": False})

    metas = m.get("metas", [])
    metas_pts = min(len([mt for mt in metas if mt.get("atual",0) > 0]) * 5, 15)
    score += metas_pts
    if metas: details.append({"item": f"Metas em andamento ({len(metas)})", "pts": metas_pts, "ok": metas_pts > 0})

    pend = [d for d in m["despesas"] if d.get("situacao") == "Pendente"]
    if not pend: score += 10; details.append({"item": "Sem contas pendentes", "pts": 10, "ok": True})
    else: details.append({"item": f"{len(pend)} contas pendentes", "pts": 0, "ok": False})

    score = min(max(score, 0), 100)
    if score >= 80: label = "Excelente 🏆"
    elif score >= 60: label = "Bom 👍"
    elif score >= 40: label = "Regular ⚠️"
    else: label = "Crítico 🚨"
    return {"score": score, "label": label, "details": details}

def detect_bank(columns: list[str]) -> str:
    cols = " ".join(c.lower() for c in columns)
    if "nubank" in cols or ("data" in cols and "descricao" in cols and "valor" in cols and "tipo" not in cols): return "nubank"
    if "agencia" in cols or "conta" in cols and "bradesco" in cols: return "bradesco"
    if "data lanc" in cols or "lançamento" in cols: return "itau"
    if "caixa" in cols or "operacao" in cols: return "caixa"
    return "inter"

def parse_csv_multi(content: str) -> tuple[list[dict], list[dict]]:
    """Tenta parsear CSV de qualquer banco suportado."""
    import csv as csv_mod
    lines = content.splitlines()

    # encontrar linha de cabeçalho
    header_idx = 0
    for i, ln in enumerate(lines):
        low = ln.lower()
        if "data" in low and any(x in low for x in ["valor","histórico","historico","descrição","descricao","lançamento","lancamento"]):
            header_idx = i; break

    csv_body = "\n".join(lines[header_idx:])
    rows_raw = []
    for sep in [";", ","]:
        try:
            r = list(csv_mod.DictReader(io.StringIO(csv_body), delimiter=sep))
            if r and len(r[0]) >= 3: rows_raw = r; break
        except: continue

    if not rows_raw: return [], []

    def norm(k): return (k or "").strip().lower() \
        .replace("ç","c").replace("ã","a").replace("é","e").replace("í","i") \
        .replace("ó","o").replace("ú","u").replace("â","a").replace("ê","e") \
        .replace("ô","o").replace(" ","_").replace(".","")

    rows = [{norm(k): (v or "").strip() for k, v in row.items()} for row in rows_raw]
    bank = detect_bank(list(rows[0].keys()))

    # mapeamento por banco
    MAPPINGS = {
        "inter":    {"data": ["data","data_lancamento"],    "desc": ["historico","descricao"],     "valor": ["valor"],       "tipo": ["tipo"]},
        "nubank":   {"data": ["data"],                      "desc": ["descricao","titulo"],         "valor": ["valor","montante"], "tipo": []},
        "bradesco": {"data": ["data","data_lancamento"],    "desc": ["historico","descricao"],     "valor": ["valor","debito","credito"], "tipo": ["tipo","db/cr"]},
        "itau":     {"data": ["data_lanc","data"],          "desc": ["historico","descricao"],     "valor": ["valor"],       "tipo": ["db/cr","tipo"]},
        "caixa":    {"data": ["data"],                      "desc": ["descricao","historico"],     "valor": ["valor","debito"], "tipo": ["operacao","tipo"]},
    }
    mp = MAPPINGS.get(bank, MAPPINGS["inter"])

    def find_col(keys):
        for k in keys:
            for col in rows[0]:
                if k in col: return col
        return None

    col_data  = find_col(mp["data"])
    col_desc  = find_col(mp["desc"])
    col_valor = find_col(mp["valor"])
    col_tipo  = find_col(mp["tipo"])

    if not col_valor: return [], []

    debitos, creditos = [], []
    for row in rows:
        val  = clean_val(row.get(col_valor,"0"))
        desc = row.get(col_desc, "") if col_desc else ""
        data = row.get(col_data, "") if col_data else ""
        tipo = row.get(col_tipo, "").lower() if col_tipo else ""
        is_deb = ("deb" in tipo or "saida" in tipo or "saída" in tipo or "-" in tipo) if col_tipo else val < 0
        item = {"descricao": desc, "valor": abs(val), "data": data, "categoria": guess_cat(desc), "banco": bank}
        if abs(val) > 0:
            (debitos if is_deb else creditos).append(item)

    return debitos, creditos

# ═══════════════════════ MODELS ══════════════════════════════

class LoginReq(BaseModel):
    user_id: str; password: str

class RegisterReq(BaseModel):
    user_id: str; name: str; password: str; color: str = "#2ecc85"

class ChangePasswordReq(BaseModel):
    old_password: str; new_password: str

class Receita(BaseModel):
    nome: str; valor: float; data_credito: str = ""

class Despesa(BaseModel):
    nome: str; categoria: str = "Essencial"; prestacao: str = ""
    valor: float; vencimento: str = ""; situacao: str = ""
    data_pgto: str = ""; obs: str = ""; ordem: int = 0; recorrente: bool = False

class Poupanca(BaseModel):
    saldo_anterior: float; aporte: float; juros: float

class Compra(BaseModel):
    item: str; valor: float; data: str = ""

class Meta(BaseModel):
    nome: str; alvo: float; atual: float = 0.0; prazo: str = ""

class GlobalMeta(BaseModel):
    nome: str; alvo: float; atual: float = 0.0
    prazo: str = ""; descricao: str = ""; auto_track: bool = False

# ═══════════════════════ HTML ════════════════════════════════

@app.get("/", response_class=HTMLResponse)
async def serve_frontend():
    if STATIC_FILE.exists():
        return HTMLResponse(content=STATIC_FILE.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>index.html não encontrado</h1>", status_code=404)

# ═══════════════════════ AUTH ROUTES ═════════════════════════

@app.get("/auth/status")
async def auth_status(session_token: Optional[str] = Cookie(default=None)):
    users = load_users()["users"]
    if not users:
        return {"setup": True, "authenticated": False, "user": None}
    uid = get_session_user(session_token)
    if uid:
        user = get_user(uid)
        return {"setup": False, "authenticated": True, "user": {"id": uid, "name": user["name"], "color": user.get("color","#2ecc85")}}
    return {"setup": False, "authenticated": False, "user": None}

@app.post("/auth/setup")
async def setup_first_user(req: RegisterReq, response: Response):
    users = load_users()
    if users["users"]:
        raise HTTPException(400, "Sistema já configurado")
    salt = secrets.token_hex(16)
    users["users"].append({"id": req.user_id, "name": req.name,
                            "salt": salt, "password_hash": hash_pw(req.password, salt),
                            "color": req.color, "role": "admin"})
    save_users(users)
    token = create_session(req.user_id)
    response.set_cookie("session_token", token, httponly=True, samesite="lax", max_age=86400*30)
    return {"ok": True, "user": {"id": req.user_id, "name": req.name, "color": req.color}}

@app.post("/auth/login")
async def login(req: LoginReq, response: Response):
    users = load_users()
    user  = next((u for u in users["users"] if u["id"] == req.user_id), None)
    if not user or hash_pw(req.password, user["salt"]) != user["password_hash"]:
        raise HTTPException(401, "Usuário ou senha incorretos")
    token = create_session(req.user_id)
    response.set_cookie("session_token", token, httponly=True, samesite="lax", max_age=86400*30)
    return {"ok": True, "user": {"id": req.user_id, "name": user["name"], "color": user.get("color","#2ecc85")}}

@app.post("/auth/logout")
async def logout(response: Response, session_token: Optional[str] = Cookie(default=None)):
    if session_token in _sessions:
        del _sessions[session_token]
    response.delete_cookie("session_token")
    return {"ok": True}

@app.post("/auth/change-password")
async def change_password(req: ChangePasswordReq, uid: str = Depends(require_auth)):
    users = load_users()
    user  = next((u for u in users["users"] if u["id"] == uid), None)
    if not user or hash_pw(req.old_password, user["salt"]) != user["password_hash"]:
        raise HTTPException(400, "Senha atual incorreta")
    user["salt"]          = secrets.token_hex(16)
    user["password_hash"] = hash_pw(req.new_password, user["salt"])
    save_users(users)
    return {"ok": True}

# ═══════════════════════ PROFILES ════════════════════════════

@app.get("/api/profiles")
async def list_profiles(uid: str = Depends(require_auth)):
    return [{"id": u["id"], "name": u["name"], "color": u.get("color","#2ecc85"), "role": u.get("role","user")}
            for u in load_users()["users"]]

@app.post("/api/profiles")
async def add_profile(req: RegisterReq, uid: str = Depends(require_auth)):
    users = load_users()
    user  = get_user(uid)
    if user.get("role") != "admin": raise HTTPException(403, "Apenas admin pode criar perfis")
    if any(u["id"] == req.user_id for u in users["users"]): raise HTTPException(400, "ID já existe")
    salt = secrets.token_hex(16)
    users["users"].append({"id": req.user_id, "name": req.name, "salt": salt,
                            "password_hash": hash_pw(req.password, salt),
                            "color": req.color, "role": "user"})
    save_users(users)
    return {"ok": True}

@app.delete("/api/profiles/{target_id}")
async def delete_profile(target_id: str, uid: str = Depends(require_auth)):
    user = get_user(uid)
    if user.get("role") != "admin": raise HTTPException(403)
    if target_id == uid: raise HTTPException(400, "Não pode excluir a si mesmo")
    users = load_users()
    users["users"] = [u for u in users["users"] if u["id"] != target_id]
    save_users(users)
    f = get_data_file(target_id)
    if f.exists(): f.unlink()
    return {"ok": True}

# ═══════════════════════ API MESES ═══════════════════════════

@app.get("/api/months")
async def get_months(uid: str = Depends(require_auth)):
    db = load_db(uid)
    out = []
    for k in sorted(k for k in db if not k.startswith("_")):
        m = migrate(db[k]); tr, td, s, *_ = month_totals(m)
        out.append({"key": k, "receitas": tr, "despesas": td, "saldo": s})
    return out

@app.get("/api/month/{key}")
async def get_month(key: str, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); save_db(uid, db)
    return db[key]

@app.put("/api/month/{key}")
async def save_month_full(key: str, payload: dict, uid: str = Depends(require_auth)):
    db = load_db(uid); db[key] = payload; save_db(uid, db)
    return {"ok": True}

# Receitas
@app.post("/api/month/{key}/receitas")
async def add_receita(key: str, r: Receita, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["receitas"].append(r.model_dump()); save_db(uid, db); return {"ok": True}

@app.put("/api/month/{key}/receitas/{idx}")
async def upd_receita(key: str, idx: int, r: Receita, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["receitas"][idx] = r.model_dump(); save_db(uid, db); return {"ok": True}

@app.delete("/api/month/{key}/receitas/{idx}")
async def del_receita(key: str, idx: int, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["receitas"].pop(idx); save_db(uid, db); return {"ok": True}

# Despesas
@app.post("/api/month/{key}/despesas")
async def add_despesa(key: str, d: Despesa, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["despesas"].append(d.model_dump()); save_db(uid, db); return {"ok": True}

@app.put("/api/month/{key}/despesas")
async def replace_despesas(key: str, despesas: list[Despesa], uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["despesas"] = [d.model_dump() for d in despesas]; save_db(uid, db); return {"ok": True}

@app.put("/api/month/{key}/despesas/{idx}")
async def upd_despesa(key: str, idx: int, d: Despesa, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["despesas"][idx] = d.model_dump(); save_db(uid, db); return {"ok": True}

@app.delete("/api/month/{key}/despesas/{idx}")
async def del_despesa(key: str, idx: int, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["despesas"].pop(idx); save_db(uid, db); return {"ok": True}

# Poupança
@app.put("/api/month/{key}/poupanca")
async def upd_poupanca(key: str, p: Poupanca, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["poupanca"] = p.model_dump(); save_db(uid, db); return {"ok": True}

# Compras
@app.post("/api/month/{key}/compras")
async def add_compra(key: str, c: Compra, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["compras"].append(c.model_dump()); save_db(uid, db); return {"ok": True}

@app.delete("/api/month/{key}/compras/{idx}")
async def del_compra(key: str, idx: int, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["compras"].pop(idx); save_db(uid, db); return {"ok": True}

# Metas
@app.post("/api/month/{key}/metas")
async def add_meta(key: str, m: Meta, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["metas"].append(m.model_dump()); save_db(uid, db); return {"ok": True}

@app.put("/api/month/{key}/metas/{idx}")
async def upd_meta(key: str, idx: int, m: Meta, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["metas"][idx] = m.model_dump(); save_db(uid, db); return {"ok": True}

@app.delete("/api/month/{key}/metas/{idx}")
async def del_meta(key: str, idx: int, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key); db[key]["metas"].pop(idx); save_db(uid, db); return {"ok": True}

# Copiar mês
@app.post("/api/month/{key}/copy-from/{src}")
async def copy_from(key: str, src: str, uid: str = Depends(require_auth)):
    db = load_db(uid)
    if src not in db: raise HTTPException(404)
    ensure_month(db, key)
    db[key]["despesas"] = [{**d,"valor":0.0,"situacao":"","data_pgto":"","obs":""} for d in db[src]["despesas"]]
    save_db(uid, db); return {"ok": True}

# ═══════════════════════ INTELIGÊNCIA FINANCEIRA ═════════════

@app.get("/api/month/{key}/score")
async def get_score(key: str, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key)
    m  = migrate(db[key]); tr, td, s, ess, pou, des = month_totals(m)
    return calc_score(m, tr, td, ess, pou, des)

@app.get("/api/month/{key}/forecast")
async def get_forecast(key: str, uid: str = Depends(require_auth)):
    db   = load_db(uid); ensure_month(db, key)
    prev = sorted([k for k in db if not k.startswith("_") and k < key])[-3:]
    if not prev: return {"available": False}

    avg_spend = sum(sum(d["valor"] for d in migrate(db[k])["despesas"]) for k in prev) / len(prev)
    avg_rec   = sum(sum(r["valor"] for r in migrate(db[k])["receitas"]) for k in prev) / len(prev)
    cur       = migrate(db[key])
    cur_spend = sum(d["valor"] for d in cur["despesas"])
    cur_rec   = sum(r["valor"] for r in cur["receitas"])

    today = date.today()
    days_in_month = cal_mod.monthrange(today.year, today.month)[1]
    day = today.day
    daily_rate  = cur_spend / max(day, 1)
    projected   = daily_rate * days_in_month
    expected_rec = cur_rec if cur_rec > 0 else avg_rec

    return {
        "available":      True,
        "current_spend":  cur_spend,
        "projected":      projected,
        "avg_historical": avg_spend,
        "expected_rec":   expected_rec,
        "projected_saldo":expected_rec - projected,
        "day":            day,
        "days_total":     days_in_month,
        "alert":          projected > avg_spend * 1.15,
    }

@app.get("/api/month/{key}/trends")
async def get_trends(key: str, uid: str = Depends(require_auth)):
    db   = load_db(uid)
    prev = sorted([k for k in db if not k.startswith("_") and k < key])[-3:]
    if len(prev) < 2: return {"available": False, "trends": []}

    def cat_totals(k):
        m = migrate(db[k])
        out = {}
        for d in m["despesas"]:
            cat = d.get("categoria","Outros")
            out[cat] = out.get(cat, 0) + d["valor"]
        return out

    trends = []
    last   = cat_totals(prev[-1])
    before = cat_totals(prev[-2])
    for cat, val in last.items():
        old = before.get(cat, 0)
        if old > 0 and val > 0:
            pct = (val - old) / old * 100
            if abs(pct) >= 10:
                trends.append({"categoria": cat, "variacao": round(pct,1),
                                "valor_atual": val, "valor_anterior": old,
                                "crescimento": pct > 0})
    trends.sort(key=lambda x: abs(x["variacao"]), reverse=True)
    return {"available": True, "trends": trends[:5], "periodo": f"{prev[-2]} → {prev[-1]}"}

@app.get("/api/month/{key}/compare")
async def compare_month(key: str, uid: str = Depends(require_auth)):
    db   = load_db(uid)
    prev = sorted([k for k in db if not k.startswith("_") and k < key])
    if not prev: return {"available": False}
    pk  = prev[-1]
    cur = migrate(db.get(key, default_month()))
    prv = migrate(db[pk])
    tr_c, td_c, s_c, *_ = month_totals(cur)
    tr_p, td_p, s_p, *_ = month_totals(prv)
    return {
        "available":   True,
        "prev_key":    pk,
        "receitas":    {"atual": tr_c, "anterior": tr_p, "diff": tr_c - tr_p},
        "despesas":    {"atual": td_c, "anterior": td_p, "diff": td_c - td_p},
        "saldo":       {"atual": s_c,  "anterior": s_p,  "diff": s_c  - s_p},
    }

@app.get("/api/month/{key}/notifications")
async def get_notifications(key: str, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key)
    m  = migrate(db[key])
    today  = date.today()
    alerts = []
    for d in m["despesas"]:
        if d.get("situacao") == "Pago" or not d.get("vencimento") or d["valor"] <= 0: continue
        venc_str = d["vencimento"]
        # tenta parsear dd/mm ou dd/mm/aaaa
        try:
            parts = venc_str.replace("-","/").split("/")
            if len(parts) >= 2:
                day = int(parts[0]); month = int(parts[1])
                year = int(parts[2]) if len(parts) > 2 else today.year
                venc = date(year, month, day)
                diff = (venc - today).days
                if diff <= 5:
                    alerts.append({"nome": d["nome"], "valor": d["valor"],
                                   "vencimento": venc_str, "dias": diff,
                                   "situacao": d.get("situacao",""),
                                   "urgente": diff <= 2})
        except: pass
    alerts.sort(key=lambda x: x["dias"])
    return {"notifications": alerts, "count": len(alerts)}

# ═══════════════════════ CALENDÁRIO ══════════════════════════

@app.get("/api/calendar/{key}")
async def get_calendar(key: str, uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key)
    m  = migrate(db[key])
    mm, yy = key.split(".")
    days_in = cal_mod.monthrange(int(yy), int(mm))[1]
    weekday_start, _ = cal_mod.monthrange(int(yy), int(mm))

    by_day: dict[int, list] = {}
    for d in m["despesas"]:
        if not d.get("vencimento") or d["valor"] <= 0: continue
        try:
            day = int(d["vencimento"].split("/")[0])
            if 1 <= day <= days_in:
                by_day.setdefault(day, []).append({
                    "nome": d["nome"][:25], "valor": d["valor"],
                    "situacao": d.get("situacao",""), "categoria": d.get("categoria","")
                })
        except: pass

    return {"weekday_start": weekday_start, "days_in_month": days_in,
            "by_day": {str(k): v for k, v in by_day.items()},
            "month": mm, "year": yy}

# ═══════════════════════ CONSOLIDADO FAMILIAR ════════════════

@app.get("/api/family/{key}")
async def family_view(key: str, uid: str = Depends(require_auth)):
    users = load_users()["users"]
    result = []
    for u in users:
        db = load_db(u["id"])
        ensure_month(db, key)
        m  = migrate(db[key])
        tr, td, s, ess, pou, des = month_totals(m)
        result.append({
            "user_id": u["id"], "name": u["name"], "color": u.get("color","#2ecc85"),
            "receitas": tr, "despesas": td, "saldo": s,
            "essencial": ess, "poupanca": pou, "desejo": des,
        })
    total_rec  = sum(r["receitas"] for r in result)
    total_desp = sum(r["despesas"] for r in result)
    return {"profiles": result, "total_receitas": total_rec,
            "total_despesas": total_desp, "saldo_familiar": total_rec - total_desp}

# ═══════════════════════ EXTRATO MULTI-BANCO ═════════════════

@app.post("/api/import-extrato")
async def import_extrato(file: UploadFile = File(...), uid: str = Depends(require_auth)):
    raw = await file.read()
    enc = chardet.detect(raw)["encoding"] or "latin-1"
    try:    content = raw.decode(enc)
    except: content = raw.decode("latin-1", errors="replace")
    debitos, creditos = parse_csv_multi(content)
    if not debitos and not creditos:
        raise HTTPException(400, "Não foi possível ler o CSV. Verifique o formato.")
    return {"debitos": debitos, "creditos": creditos,
            "banco_detectado": debitos[0].get("banco","?") if debitos else "?"}

@app.post("/api/month/{key}/check-duplicates")
async def check_duplicates(key: str, items: list[dict], uid: str = Depends(require_auth)):
    db = load_db(uid); ensure_month(db, key)
    m  = migrate(db[key])
    existing = [(d["nome"].lower()[:20], d["valor"]) for d in m["despesas"] if d["valor"] > 0]
    existing += [(c["item"].lower()[:20], c["valor"]) for c in m["compras"]]
    dups = []
    for item in items:
        desc = item.get("descricao","").lower()[:20]
        val  = float(item.get("valor", 0))
        is_dup = any(abs(v - val) < 0.01 and (desc[:10] in n or n[:10] in desc)
                     for n, v in existing)
        dups.append({**item, "possivel_duplicata": is_dup})
    return dups

# ═══════════════════════ BACKUPS ══════════════════════════════

@app.get("/api/backups")
async def list_backups(uid: str = Depends(require_auth)):
    baks = sorted(BACKUP_DIR.glob(f"backup_{uid}_*.json"), reverse=True)[:20]
    return [{"filename": b.name, "size_kb": round(b.stat().st_size/1024,1),
             "date": b.stat().st_mtime} for b in baks]

@app.post("/api/backups/restore/{filename}")
async def restore_backup(filename: str, uid: str = Depends(require_auth)):
    f = BACKUP_DIR / filename
    if not f.exists() or uid not in filename: raise HTTPException(404)
    data = json.loads(f.read_text(encoding="utf-8"))
    save_db(uid, data); return {"ok": True}

@app.get("/api/backups/download/{filename}")
async def download_backup(filename: str, uid: str = Depends(require_auth)):
    f = BACKUP_DIR / filename
    if not f.exists() or uid not in filename: raise HTTPException(404)
    return StreamingResponse(io.BytesIO(f.read_bytes()), media_type="application/json",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})

# ═══════════════════════ ANÁLISE PERÍODO / ANUAL ═════════════

@app.get("/api/period-analysis")
async def period_analysis(start: str = Query(...), end: str = Query(...),
                          uid: str = Depends(require_auth)):
    db = load_db(uid)
    keys = sorted(k for k in db if not k.startswith("_") and start <= k <= end)
    if not keys: raise HTTPException(404)
    by_cat: dict[str,float] = {}; by_nome: dict[str,float] = {}
    total_rec = 0.0; total_desp = 0.0
    for k in keys:
        mv = migrate(db[k]); tr, td, *_ = month_totals(mv)
        total_rec += tr; total_desp += td
        for d in mv["despesas"]:
            if d["valor"] > 0:
                by_cat[d.get("categoria","Outros")] = by_cat.get(d.get("categoria","Outros"),0)+d["valor"]
                by_nome[d["nome"]] = by_nome.get(d["nome"],0)+d["valor"]
    return {"meses": keys, "total_receitas": total_rec, "total_despesas": total_desp,
            "saldo": total_rec-total_desp, "por_categoria": by_cat,
            "top_despesas": [{"nome":n,"valor":v} for n,v in sorted(by_nome.items(),key=lambda x:x[1],reverse=True)[:10]]}

@app.get("/api/annual/{ano}")
async def get_annual(ano: str, uid: str = Depends(require_auth)):
    db = load_db(uid); rows = []
    for k, m in sorted((k,v) for k,v in db.items() if k.endswith(f".{ano}") and not k.startswith("_")):
        m = migrate(m); tr,td,s,ess,pou,des = month_totals(m)
        p = m["poupanca"]
        rows.append({"mes":k,"receitas":tr,"despesas":td,"saldo":s,"essencial":ess,
                     "poupanca":pou,"desejo":des,"poupanca_acum":p["saldo_anterior"]+p["aporte"]+p["juros"]})
    return rows

# ═══════════════════════ METAS GLOBAIS ═══════════════════════

@app.get("/api/global-metas")
async def get_global_metas(uid: str = Depends(require_auth)):
    db = load_db(uid); metas = db.get("_global_metas",[])
    for meta in metas:
        if meta.get("auto_track"):
            meta["atual_calculado"] = sum(
                v["poupanca"].get("aporte",0) for k,v in db.items()
                if not k.startswith("_") and isinstance(v,dict) and "poupanca" in v)
    return metas

@app.post("/api/global-metas")
async def add_global_meta(m: GlobalMeta, uid: str = Depends(require_auth)):
    db = load_db(uid); db.setdefault("_global_metas",[]).append(m.model_dump()); save_db(uid,db); return {"ok":True}

@app.put("/api/global-metas/{idx}")
async def upd_global_meta(idx: int, m: GlobalMeta, uid: str = Depends(require_auth)):
    db = load_db(uid); db["_global_metas"][idx]=m.model_dump(); save_db(uid,db); return {"ok":True}

@app.delete("/api/global-metas/{idx}")
async def del_global_meta(idx: int, uid: str = Depends(require_auth)):
    db = load_db(uid); db["_global_metas"].pop(idx); save_db(uid,db); return {"ok":True}

# ═══════════════════════ EXPORT PDF / EXCEL ══════════════════

@app.get("/api/month/{key}/pdf")
async def month_pdf(key: str, uid: str = Depends(require_auth)):
    try: from fpdf import FPDF
    except: raise HTTPException(500,"fpdf2 não instalado")
    db = load_db(uid)
    if key not in db: raise HTTPException(404)
    m  = migrate(db[key]); tr,td,s,ess,pou,des = month_totals(m)
    pdf = FPDF(); pdf.add_page(); pdf.set_auto_page_break(True,15)
    pdf.set_fill_color(26,71,42); pdf.rect(0,0,210,28,"F")
    pdf.set_font("Helvetica","B",16); pdf.set_text_color(255,255,255)
    pdf.cell(0,10,"",ln=True); pdf.cell(0,10,f"Orcamento Mensal  {key}",align="C",ln=True)
    pdf.set_text_color(0,0,0); pdf.ln(4)

    score_data = calc_score(m,tr,td,ess,pou,des)
    pdf.set_font("Helvetica","B",11)
    pdf.cell(0,7,f"Score Financeiro: {score_data['score']}/100 — {score_data['label']}",ln=True)
    pdf.ln(3)
    for label,val,r,g,b in [("Receitas",tr,39,174,96),("Despesas",td,231,76,60),("Saldo",s,39,174,96 if s>=0 else 231,76,60)]:
        pdf.set_fill_color(r,g,b); pdf.set_text_color(255,255,255); pdf.set_font("Helvetica","B",9)
        pdf.cell(60,8,f"  {label}",fill=True); pdf.set_text_color(50,50,50); pdf.set_fill_color(245,245,245)
        pdf.cell(55,8,f"  {brl(val)}",fill=True,ln=True)
    pdf.ln(3)

    def sec(t):
        pdf.set_font("Helvetica","B",11); pdf.set_text_color(26,71,42); pdf.cell(0,7,t,ln=True)
        pdf.set_text_color(0,0,0); pdf.set_draw_color(200,200,200); pdf.line(10,pdf.get_y(),200,pdf.get_y()); pdf.ln(2)

    sec("Receitas")
    pdf.set_font("Helvetica","",9)
    for r in m["receitas"]:
        if r["valor"]>0: pdf.cell(80,5,f"  {r['nome']}"); pdf.cell(40,5,brl(r["valor"]),ln=True)

    sec("Despesas")
    pdf.set_font("Helvetica","",8)
    CAT_C={"Essencial":(41,128,185),"Poupança":(39,174,96),"Desejo":(230,126,34)}
    for d in m["despesas"]:
        if d["valor"]>0:
            r2,g2,b2=CAT_C.get(d.get("categoria","Essencial"),(120,120,120))
            pdf.set_text_color(r2,g2,b2); pdf.cell(8,5,d.get("categoria","?")[0])
            pdf.set_text_color(0,0,0); pdf.cell(90,5,d["nome"])
            pdf.cell(28,5,d.get("situacao",""))
            pdf.cell(30,5,brl(d["valor"]),ln=True)

    pdf.ln(3); pdf.set_font("Helvetica","I",7); pdf.set_text_color(150,150,150)
    pdf.cell(0,5,f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",align="C")
    buf=io.BytesIO(pdf.output())
    return StreamingResponse(buf,media_type="application/pdf",
                             headers={"Content-Disposition":f"attachment; filename=orcamento_{key}.pdf"})

@app.get("/api/month/{key}/excel")
async def month_excel(key: str, uid: str = Depends(require_auth)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except: raise HTTPException(500,"openpyxl não instalado")
    db=load_db(uid)
    if key not in db: raise HTTPException(404)
    m=migrate(db[key]); tr,td,s,ess,pou,des=month_totals(m)
    wb=Workbook(); ws=wb.active; ws.title="Resumo"
    for col,w in zip('ABCDE',[28,16,14,14,14]): ws.column_dimensions[col].width=w
    ws['A1']=f"Orçamento Mensal — {key}"; ws['A1'].font=Font(bold=True,size=14)
    for i,(lbl,val) in enumerate([("Receitas",tr),("Despesas",td),("Saldo",s),("Essenciais",ess),("Poupança",pou),("Desejos",des)],3):
        ws.cell(i,1,lbl).font=Font(bold=True); c=ws.cell(i,2,val); c.number_format='R$ #,##0.00'
    ws2=wb.create_sheet("Receitas"); ws2.append(["Nome","Valor","Data"])
    for r in m["receitas"]: ws2.append([r["nome"],r["valor"],r.get("data_credito","")]); ws2.cell(ws2.max_row,2).number_format='R$ #,##0.00'
    ws3=wb.create_sheet("Despesas"); ws3.append(["Nome","Categoria","Valor","Vencimento","Situação","Recorrente"])
    for d in m["despesas"]: ws3.append([d["nome"],d.get("categoria",""),d["valor"],d.get("vencimento",""),d.get("situacao",""),"Sim" if d.get("recorrente") else "Não"]); ws3.cell(ws3.max_row,3).number_format='R$ #,##0.00'
    ws4=wb.create_sheet("Compras"); ws4.append(["Item","Valor","Data"])
    for c in m["compras"]: ws4.append([c["item"],c["valor"],c.get("data","")]); ws4.cell(ws4.max_row,2).number_format='R$ #,##0.00'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition":f"attachment; filename=orcamento_{key}.xlsx"})

@app.get("/api/annual/{ano}/excel")
async def annual_excel(ano: str, uid: str = Depends(require_auth)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except: raise HTTPException(500)
    db=load_db(uid); rows=[]
    for k,mv in sorted((k,v) for k,v in db.items() if k.endswith(f".{ano}") and not k.startswith("_")):
        mv=migrate(mv); tr,td,s,ess,pou,des=month_totals(mv); p=mv["poupanca"]
        rows.append({"mes":k,"receitas":tr,"despesas":td,"saldo":s,"essencial":ess,"poupanca":pou,"desejo":des,"poupanca_acum":p["saldo_anterior"]+p["aporte"]+p["juros"]})
    if not rows: raise HTTPException(404)
    wb=Workbook(); ws=wb.active; ws.title=f"Anual {ano}"
    ws.append(["Mês","Receitas","Despesas","Saldo","Essencial","Poupança","Desejo","Poupança Acum."])
    for r in rows:
        ws.append([r["mes"],r["receitas"],r["despesas"],r["saldo"],r["essencial"],r["poupanca"],r["desejo"],r["poupanca_acum"]])
        for ci in range(2,9): ws.cell(ws.max_row,ci).number_format='R$ #,##0.00'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition":f"attachment; filename=anual_{ano}.xlsx"})

@app.get("/api/annual/{ano}/pdf")
async def annual_pdf(ano: str, uid: str = Depends(require_auth)):
    try: from fpdf import FPDF
    except: raise HTTPException(500)
    db=load_db(uid); rows=[]
    for k,mv in sorted((k,v) for k,v in db.items() if k.endswith(f".{ano}") and not k.startswith("_")):
        mv=migrate(mv); tr,td,s,ess,pou,des=month_totals(mv); p=mv["poupanca"]
        rows.append({"mes":k,"receitas":tr,"despesas":td,"saldo":s,"poupanca_acum":p["saldo_anterior"]+p["aporte"]+p["juros"],"despesas_list":mv["despesas"]})
    if not rows: raise HTTPException(404)
    pdf=FPDF(); pdf.add_page(); pdf.set_auto_page_break(True,15)
    pdf.set_fill_color(26,71,42); pdf.rect(0,0,210,35,"F")
    pdf.set_font("Helvetica","B",18); pdf.set_text_color(255,255,255)
    pdf.cell(0,15,"",ln=True); pdf.cell(0,12,f"Relatorio Anual  {ano}",align="C",ln=True)
    pdf.set_text_color(0,0,0); pdf.ln(5)
    tot_r=sum(r["receitas"] for r in rows); tot_d=sum(r["despesas"] for r in rows)
    pdf.set_font("Helvetica","B",10)
    pdf.cell(0,7,f"Receita Anual: {brl(tot_r)}   Despesa Anual: {brl(tot_d)}   Saldo: {brl(tot_r-tot_d)}",ln=True)
    pdf.ln(3)
    pdf.set_font("Helvetica","B",9); pdf.set_fill_color(26,71,42); pdf.set_text_color(255,255,255)
    for h,w in [("Mes",20),("Receitas",38),("Despesas",38),("Saldo",38),("Poupanca",38)]: pdf.cell(w,6,h,fill=True)
    pdf.ln(); pdf.set_text_color(0,0,0)
    for r in rows:
        pdf.set_font("Helvetica","",8); pdf.cell(20,5,r["mes"])
        for v,c in [(r["receitas"],(39,174,96)),(r["despesas"],(231,76,60)),(r["saldo"],(39,174,96) if r["saldo"]>=0 else (231,76,60)),(r["poupanca_acum"],(41,128,185))]:
            pdf.set_text_color(*c); pdf.cell(38,5,brl(v))
        pdf.set_text_color(0,0,0); pdf.ln()
    pdf.ln(3); pdf.set_font("Helvetica","I",7); pdf.set_text_color(150,150,150)
    pdf.cell(0,5,f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",align="C")
    buf=io.BytesIO(pdf.output())
    return StreamingResponse(buf,media_type="application/pdf",
                             headers={"Content-Disposition":f"attachment; filename=relatorio_anual_{ano}.pdf"})

# ═══════════════════════ MAIN ════════════════════════════════

if __name__ == "__main__":
    print("\n" + "═"*52)
    print("  💰 ORÇAMENTO MENSAL v4.0")
    print("  📡 http://localhost:8001")
    print("  📚 Docs: http://localhost:8001/docs")
    print("═"*52 + "\n")
    uvicorn.run("app:app", host="0.0.0.0", port=8001, reload=False)


# ═══════════════════════ OPEN FINANCE / PLUGGY ═══════════════
"""
Pluggy é uma fintech brasileira registrada no Bacen como TPP do Open Finance.
Suporta: Nubank, Itaú, Bradesco, Caixa, BB, Santander, XP, Inter e +50 bancos.
Sandbox gratuito: https://pluggy.ai → criar conta → Dashboard → Credenciais
"""

PLUGGY_BASE = "https://api.pluggy.ai"

def _pluggy_check():
    if not REQUESTS_OK:
        raise HTTPException(500, "Instale requests: pip install requests")

def pluggy_apikey(client_id: str, client_secret: str) -> str:
    _pluggy_check()
    try:
        r = req_lib.post(f"{PLUGGY_BASE}/auth",
                         json={"clientId": client_id, "clientSecret": client_secret},
                         timeout=12)
        r.raise_for_status()
        return r.json()["apiKey"]
    except Exception as e:
        raise HTTPException(400, f"Credenciais inválidas ou sem conexão: {e}")

def pluggy_req(method: str, path: str, api_key: str, **kwargs):
    _pluggy_check()
    r = getattr(req_lib, method)(
        f"{PLUGGY_BASE}{path}",
        headers={"X-API-KEY": api_key},
        timeout=15, **kwargs)
    if not r.ok:
        raise HTTPException(r.status_code, r.text[:300])
    return r.json()

class PluggyConfig(BaseModel):
    client_id: str; client_secret: str


@app.get("/pluggy-connect.js")
async def serve_pluggy_sdk():
    """Serve o SDK Pluggy localmente se o arquivo existir."""
    f = Path("pluggy-connect.js")
    if f.exists():
        from fastapi.responses import FileResponse
        return FileResponse(f, media_type="application/javascript")
    raise HTTPException(404, "pluggy-connect.js nao encontrado. Baixe em: https://cdn.jsdelivr.net/npm/pluggy-connect-sdk/dist/pluggy-connect.js")

@app.get("/api/openfinance/config-status")
async def of_status(uid: str = Depends(require_auth)):
    db = load_db(uid); cfg = db.get("_pluggy")
    return {"configured": bool(cfg),
            "client_id":  cfg.get("client_id","") if cfg else "",
            "items":      db.get("_pluggy_items",[])}

@app.post("/api/openfinance/config")
async def of_save_config(config: PluggyConfig, uid: str = Depends(require_auth)):
    pluggy_apikey(config.client_id, config.client_secret)   # valida antes
    db = load_db(uid)
    db["_pluggy"] = {"client_id": config.client_id, "client_secret": config.client_secret}
    save_db(uid, db)
    return {"ok": True}



@app.get("/api/openfinance/debug-token")
async def of_debug_token():
    """Debug: mostra o token bruto retornado pelo Pluggy."""
    # Tenta carregar qualquer users.json para pegar config
    import glob
    cfg = None
    for f in glob.glob("data_*.json"):
        import json as _json
        data = _json.loads(open(f).read())
        if data.get("_pluggy"):
            cfg = data["_pluggy"]; break
    if not cfg: raise HTTPException(400, "Sem config Pluggy em nenhum perfil")
    key = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
    r   = req_lib.post(f"{PLUGGY_BASE}/connect_token",
                       headers={"X-API-KEY": key}, json={}, timeout=15)
    return {"status_code": r.status_code, "raw": r.json(),
            "api_key_start": key[:30]+"..."}

@app.post("/api/openfinance/connect-token")
async def of_connect_token(uid: str = Depends(require_auth)):
    """Gera um connectToken Pluggy para abrir o widget de conexão bancária."""
    db = load_db(uid); cfg = db.get("_pluggy")
    if not cfg: raise HTTPException(400, "Configure Pluggy primeiro")
    key = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
    r   = pluggy_req("post", "/connect_token", key, json={})
    # Pluggy pode retornar accessToken ou connectToken dependendo da versão
    token = r.get("accessToken") or r.get("connectToken") or r.get("access_token") or r.get("connect_token") or ""
    if not token:
        raise HTTPException(500, f"Token não encontrado na resposta Pluggy: {list(r.keys())}")
    return {"connect_token": token}

@app.post("/api/openfinance/save-item")
async def of_save_item(body: dict, uid: str = Depends(require_auth)):
    """Salva o itemId retornado pelo widget Pluggy Connect após conexão bem-sucedida."""
    db    = load_db(uid); cfg = db.get("_pluggy")
    if not cfg: raise HTTPException(400)
    key   = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
    item_id = body["item_id"]
    # busca detalhes do item
    try:
        r = pluggy_req("get", f"/items/{item_id}", key)
    except:
        r = {"id": item_id, "connector": {}}
    items = db.get("_pluggy_items", [])
    if not any(i["id"] == item_id for i in items):
        items.append({"id": item_id, "connector": r.get("connector", {}),
                      "status": r.get("status","UPDATED"),
                      "added_at": datetime.now().isoformat()})
    db["_pluggy_items"] = items; save_db(uid, db)
    return {"ok": True}

@app.delete("/api/openfinance/config")
async def of_delete_config(uid: str = Depends(require_auth)):
    db = load_db(uid); db.pop("_pluggy", None); db.pop("_pluggy_items", None)
    save_db(uid, db); return {"ok": True}

@app.get("/api/openfinance/connectors")
async def of_connectors(search: str = Query(default=""), uid: str = Depends(require_auth)):
    db = load_db(uid); cfg = db.get("_pluggy")
    if not cfg: raise HTTPException(400, "Configure as credenciais Pluggy primeiro")
    key = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
    r   = pluggy_req("get", "/connectors", key)
    out = []
    for c in r.get("results", []):
        if search and search.lower() not in c["name"].lower(): continue
        out.append({"id": c["id"], "name": c["name"], "type": c.get("type",""),
                    "logo": c.get("imageUrl",""), "health": c.get("health",{}).get("status",""),
                    "sandbox": c.get("isSandbox", False), "products": c.get("products",[]),
                    "credentials": c.get("credentials",[])})
    return sorted(out, key=lambda x: x["name"])

@app.get("/api/openfinance/items")
async def of_items(uid: str = Depends(require_auth)):
    return load_db(uid).get("_pluggy_items", [])

@app.post("/api/openfinance/connect")
async def of_connect(body: dict, uid: str = Depends(require_auth)):
    db = load_db(uid); cfg = db.get("_pluggy")
    if not cfg: raise HTTPException(400, "Configure Pluggy primeiro")
    key = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
    r   = pluggy_req("post", "/items", key,
                     json={"connectorId": int(body["connector_id"]),
                           "parameters":  body.get("parameters", {})})
    items = db.get("_pluggy_items", [])
    if not any(i["id"] == r["id"] for i in items):
        items.append({"id": r["id"], "connector": r.get("connector",{}),
                      "status": r.get("status",""), "added_at": datetime.now().isoformat()})
    db["_pluggy_items"] = items; save_db(uid, db)
    return r

@app.get("/api/openfinance/items/{item_id}/status")
async def of_item_status(item_id: str, uid: str = Depends(require_auth)):
    db = load_db(uid); cfg = db.get("_pluggy")
    if not cfg: raise HTTPException(400)
    key = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
    r   = pluggy_req("get", f"/items/{item_id}", key)
    for item in db.get("_pluggy_items",[]):
        if item["id"] == item_id: item["status"] = r.get("status","")
    save_db(uid, db); return r

@app.delete("/api/openfinance/items/{item_id}")
async def of_delete_item(item_id: str, uid: str = Depends(require_auth)):
    db = load_db(uid); cfg = db.get("_pluggy")
    if cfg:
        try:
            key = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
            pluggy_req("delete", f"/items/{item_id}", key)
        except: pass
    db["_pluggy_items"] = [i for i in db.get("_pluggy_items",[]) if i["id"] != item_id]
    save_db(uid, db); return {"ok": True}

@app.get("/api/openfinance/items/{item_id}/accounts")
async def of_accounts(item_id: str, uid: str = Depends(require_auth)):
    db = load_db(uid); cfg = db.get("_pluggy")
    if not cfg: raise HTTPException(400)
    key = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
    return pluggy_req("get", f"/accounts?itemId={item_id}", key)

@app.get("/api/openfinance/accounts/{account_id}/transactions")
async def of_transactions(account_id: str,
                          date_from: str = Query(default=""),
                          date_to:   str = Query(default=""),
                          uid: str = Depends(require_auth)):
    db = load_db(uid); cfg = db.get("_pluggy")
    if not cfg: raise HTTPException(400)
    key    = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
    params = {"accountId": account_id, "pageSize": 200}
    if date_from: params["from"] = date_from
    if date_to:   params["to"]   = date_to
    return pluggy_req("get", "/transactions", key, params=params)

@app.post("/api/openfinance/accounts/{account_id}/import/{month_key}")
async def of_import(account_id: str, month_key: str,
                    body: dict, uid: str = Depends(require_auth)):
    import calendar as cal2
    db  = load_db(uid); ensure_month(db, month_key)
    cfg = db.get("_pluggy")
    if not cfg: raise HTTPException(400)
    api_key = pluggy_apikey(cfg["client_id"], cfg["client_secret"])
    mm, yy  = month_key.split(".")
    days    = cal2.monthrange(int(yy), int(mm))[1]
    params  = {"accountId": account_id, "pageSize": 200,
               "from": f"{yy}-{mm}-01", "to": f"{yy}-{mm}-{days:02d}"}
    txns    = pluggy_req("get", "/transactions", api_key, params=params).get("results", [])
    sel_ids = set(body.get("ids", []))
    if sel_ids: txns = [t for t in txns if t["id"] in sel_ids]

    imp_d = imp_c = 0
    for t in txns:
        val  = abs(float(t.get("amount", 0)))
        if val <= 0: continue
        desc     = (t.get("description") or t.get("merchant",{}).get("name","") or "Transação").strip()
        date_str = (t.get("date","")[:10]).replace("-","/")
        is_debit = float(t.get("amount",0)) < 0 or t.get("type","").upper() in ("DEBIT","PAYMENT","PIX_DEBIT","TED","DOC","FEE")
        if is_debit:
            db[month_key]["despesas"].append({
                "nome": desc[:60], "categoria": guess_cat(desc), "prestacao": "",
                "valor": val, "vencimento": date_str, "situacao": "Pago",
                "data_pgto": date_str, "obs": f"Pluggy·{t.get('id','')[:12]}",
                "ordem": 999, "recorrente": False})
            imp_d += 1
        else:
            db[month_key]["receitas"].append({"nome": desc[:50], "valor": val, "data_credito": date_str})
            imp_c += 1

    save_db(uid, db)
    return {"ok": True, "despesas": imp_d, "receitas": imp_c, "total": imp_d + imp_c}

# ═══════════════════════ IMPORTAR PDF DE FATURA ══════════════

def clean_brl_val(s: str) -> float:
    s = str(s or "0").strip().replace("R$","").replace("+","").replace(" ","")
    neg = s.startswith("-"); s = s.lstrip("-")
    if re.search(r"\d\.\d{3},\d{2}", s): s = s.replace(".","").replace(",",".")
    else:                                  s = s.replace(",",".")
    try:    v = float(s); return -v if neg else v
    except: return 0.0

def _detect_bank(text: str) -> str:
    low = text.lower()
    if "banco inter" in low or "bancointer" in low or "bco inter" in low: return "Inter"
    if "nubank" in low or "nu pagamentos" in low: return "Nubank"
    if "bradesco" in low:                         return "Bradesco"
    if "itau" in low or "itaú" in low:            return "Itaú"
    if "santander" in low:                        return "Santander"
    if "caixa" in low:                            return "Caixa"
    if "c6 bank" in low:                          return "C6 Bank"
    return "Banco"

SKIP_WORDS = ["total cart","pagamento minimo","pagamento mínimo","saldo total",
              "valor total","fatura atual","despesas do mes","despesas do mês",
              "limite de","próxima fatura","proxima fatura","encargos financeiros",
              "data de corte","data vencimento","vencimento","subtotal"]

def _is_skip(desc: str) -> bool:
    low = desc.lower()
    return any(s in low for s in SKIP_WORDS) or len(desc) < 3

def _parse_inter(text: str) -> list:
    """Parser para fatura Inter: '28 de mar. 2026 DESCRIÇÃO - R$ 15,71'"""
    PAT = re.compile(
        r"(\d{2}\s+de\s+\w+\.?\s+\d{4})\s+(.+?)\s+-\s+([+]?\s*R?\$?\s*[\d\.]+,\d{2})",
        re.IGNORECASE
    )
    results = []; seen = set()
    for m in PAT.finditer(text):
        data, desc, val_s = m.group(1).strip(), m.group(2).strip(), m.group(3)
        if _is_skip(desc): continue
        val = abs(clean_brl_val(val_s))
        if val <= 0 or val > 200000: continue
        is_cred = "+" in val_s
        key = f"{data[:8]}|{desc[:15]}|{round(val,2)}"
        if key in seen: continue
        seen.add(key)
        results.append({"data": data, "descricao": desc[:60], "valor": val,
                        "is_credito": is_cred, "categoria": guess_cat(desc)})
    return results

def _parse_generic(text: str) -> list:
    PATS = [
        r"^(\d{2}/\d{2}/\d{4})\s+(.{3,60}?)\s+R?\$?\s*([\d\.]+,\d{2})\s*$",
        r"^(\d{2}/\d{2})\s+(.{3,60}?)\s+R?\$?\s*([\d\.]+,\d{2})\s*$",
        r"^(\d{2}\s+(?:jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)\.?)\s+(.{3,60}?)\s+R?\$?\s*([\d\.]+,\d{2})\s*$",
    ]
    results = []; seen = set()
    for line in text.splitlines():
        line = line.strip()
        if len(line) < 8: continue
        for pat in PATS:
            m = re.match(pat, line, re.IGNORECASE)
            if not m: continue
            data, desc, val_s = m.group(1), m.group(2).strip(), m.group(3)
            if _is_skip(desc): continue
            val = clean_brl_val(val_s)
            if val <= 0 or val > 50000: continue
            key = f"{data[:8]}|{desc[:15]}|{round(val,2)}"
            if key in seen: continue
            seen.add(key)
            results.append({"data": data, "descricao": desc[:60], "valor": val,
                            "is_credito": False, "categoria": guess_cat(desc)})
            break
    return results

def _extract_text(pdf_bytes: bytes, password: str) -> str:
    text = ""
    # pdfplumber (melhor para tabelas)
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(pdf_bytes), password=password or None) as pdf:
            for page in pdf.pages:
                try:
                    for table in (page.extract_tables() or []):
                        for row in table:
                            text += " ".join(str(c or "").strip() for c in row if c) + "\n"
                    t = page.extract_text()
                    if t: text += t + "\n"
                except: pass
    except Exception:
        pass
    # pypdf fallback
    if not text.strip():
        try:
            import pypdf
            reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
            if reader.is_encrypted:
                for pw in [password, password.replace("-","").replace(".",""), ""]:
                    try:
                        if reader.decrypt(pw or ""): break
                    except: pass
            for page in reader.pages:
                try: text += page.extract_text() + "\n"
                except: pass
        except Exception as e:
            raise ValueError(f"Não foi possível abrir o PDF: {e}")
    return text

def parse_pdf_invoice(pdf_bytes: bytes, password: str = "") -> dict:
    text  = _extract_text(pdf_bytes, password)
    if not text.strip():
        raise ValueError("Não foi possível extrair texto. PDF pode ser imagem escaneada — tente o CSV.")
    banco = _detect_bank(text)
    raw   = _parse_inter(text) if banco == "Inter" else _parse_generic(text)
    if not raw: raw = _parse_inter(text) + _parse_generic(text)  # tenta ambos
    # deduplica
    seen2 = set(); final = []
    for r in raw:
        k = f"{r['data'][:8]}|{r['descricao'][:15]}|{round(r['valor'],2)}"
        if k not in seen2: seen2.add(k); final.append(r)
    debitos  = [{k:v for k,v in r.items() if k!="is_credito"} for r in final if not r["is_credito"]]
    creditos = [{k:v for k,v in r.items() if k!="is_credito"} for r in final if r["is_credito"]]
    return {"debitos": debitos, "creditos": creditos, "banco": banco, "chars": len(text)}

@app.post("/api/import-pdf")
async def import_pdf_invoice(
    file:     UploadFile = File(...),
    password: str        = Query(default=""),
    uid: str = Depends(require_auth)
):
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Arquivo vazio")
    try:
        result = parse_pdf_invoice(raw, password)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(500, f"Erro ao processar PDF: {e}")

    if not result["debitos"] and not result["creditos"]:
        raise HTTPException(422, f"Nenhuma transação encontrada no PDF do {result['banco']}. "
                                 f"O PDF pode estar em formato de imagem ou o layout não é reconhecido. "
                                 f"Tente exportar o extrato em CSV.")
    return result
